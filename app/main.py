from __future__ import annotations

import asyncio
import os
import time
import io
from typing import List

from fastapi import Depends, FastAPI, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session, joinedload
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from . import crud, models, schemas
from .background_jobs import experiment_runner
from .database import Base, engine, get_db, SessionLocal

from .websocket import connection_manager

app = FastAPI(title="ConflictAwareLab", version="0.1.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def on_startup() -> None:
    Base.metadata.create_all(bind=engine)
    
    # Set the main event loop for the curriculum orchestrator
    # from .background_jobs import curriculum_orchestrator
    # loop = asyncio.get_event_loop()
    # curriculum_orchestrator.main_loop = loop
    # print(f"[Startup] Set main event loop for curriculum orchestrator: {type(loop)}")


# ------------------------------ Helper Functions ------------------------------ #


def _get_or_404(db: Session, model, entity_id: int):
    entity = crud.get_entity(db, model, entity_id)
    if not entity:
        raise HTTPException(status_code=404, detail="Item not found")
    return entity


# ==================== ENVIRONMENTS ====================
@app.post("/environments", response_model=schemas.Environment, status_code=201)
def create_environment(
    payload: schemas.EnvironmentCreate, db: Session = Depends(get_db)
):
    """Create a new environment configuration."""
    return crud.create_entity(db, models.Environment, payload)


@app.get("/environments", response_model=List[schemas.Environment])
def list_environments(db: Session = Depends(get_db)):
    """List all environment configurations."""
    return crud.list_entities(db, models.Environment)


@app.get("/environments/{environment_id}", response_model=schemas.Environment)
def get_environment(environment_id: int, db: Session = Depends(get_db)):
    """Get a specific environment by ID."""
    return _get_or_404(db, models.Environment, environment_id)


@app.put("/environments/{environment_id}", response_model=schemas.Environment)
def update_environment(
    environment_id: int,
    payload: schemas.EnvironmentUpdate,
    db: Session = Depends(get_db),
):
    """Update an environment. Fails if environment is locked."""
    env = _get_or_404(db, models.Environment, environment_id)
    crud.check_entity_locked(env)
    return crud.update_entity(db, env, payload)


@app.delete("/environments/{environment_id}")
def delete_environment(environment_id: int, db: Session = Depends(get_db)):
    """Delete an environment. Fails if used in any experiment."""
    env = _get_or_404(db, models.Environment, environment_id)
    crud.check_experiments_using_entity(db, 'environment', environment_id)
    crud.delete_entity(db, env)
    return {"status": "deleted"}


# ==================== AGENTS ====================
@app.post("/agents", response_model=schemas.Agent, status_code=201)
def create_agent(payload: schemas.AgentCreate, db: Session = Depends(get_db)):
    """Create a new agent configuration."""
    return crud.create_entity(db, models.Agent, payload)


@app.get("/agents", response_model=List[schemas.Agent])
def list_agents(db: Session = Depends(get_db)):
    """List all agent configurations."""
    return crud.list_entities(db, models.Agent)


@app.get("/agents/{agent_id}", response_model=schemas.Agent)
def get_agent(agent_id: int, db: Session = Depends(get_db)):
    """Get a specific agent by ID."""
    return _get_or_404(db, models.Agent, agent_id)


@app.put("/agents/{agent_id}", response_model=schemas.Agent)
def update_agent(
    agent_id: int, payload: schemas.AgentUpdate, db: Session = Depends(get_db)
):
    """Update an agent. Fails if agent is locked."""
    agent = _get_or_404(db, models.Agent, agent_id)
    crud.check_entity_locked(agent)
    return crud.update_entity(db, agent, payload)


@app.delete("/agents/{agent_id}")
def delete_agent(agent_id: int, db: Session = Depends(get_db)):
    """Delete an agent. Fails if used in any experiment."""
    agent = _get_or_404(db, models.Agent, agent_id)
    crud.check_experiments_using_entity(db, 'agent', agent_id)
    crud.delete_entity(db, agent)
    return {"status": "deleted"}


# ==================== RESIDUAL CONNECTORS ====================
@app.post("/residual-connectors", response_model=schemas.ResidualConnector, status_code=201)
def create_residual_connector(
    payload: schemas.ResidualConnectorCreate, db: Session = Depends(get_db)
):
    """Create a new residual connector (safety corrector layer)."""
    # Verify parent agent exists
    _get_or_404(db, models.Agent, payload.parent_agent_id)
    return crud.create_entity(db, models.ResidualConnector, payload)


@app.get("/residual-connectors", response_model=List[schemas.ResidualConnector])
def list_residual_connectors(db: Session = Depends(get_db)):
    """List all residual connectors."""
    return db.query(models.ResidualConnector).options(
        joinedload(models.ResidualConnector.parent_agent)
    ).all()


@app.get("/residual-connectors/{connector_id}", response_model=schemas.ResidualConnector)
def get_residual_connector(connector_id: int, db: Session = Depends(get_db)):
    """Get a specific residual connector by ID."""
    connector = db.query(models.ResidualConnector).options(
        joinedload(models.ResidualConnector.parent_agent)
    ).filter(models.ResidualConnector.id == connector_id).first()
    if not connector:
        raise HTTPException(status_code=404, detail="Residual connector not found")
    return connector


@app.put("/residual-connectors/{connector_id}", response_model=schemas.ResidualConnector)
def update_residual_connector(
    connector_id: int,
    payload: schemas.ResidualConnectorUpdate,
    db: Session = Depends(get_db)
):
    """Update a residual connector. Fails if connector is locked."""
    connector = _get_or_404(db, models.ResidualConnector, connector_id)
    crud.check_entity_locked(connector)
    return crud.update_entity(db, connector, payload)


@app.delete("/residual-connectors/{connector_id}")
def delete_residual_connector(connector_id: int, db: Session = Depends(get_db)):
    """Delete a residual connector."""
    connector = _get_or_404(db, models.ResidualConnector, connector_id)
    crud.check_entity_locked(connector)
    crud.delete_entity(db, connector)
    return {"status": "deleted"}


# ==================== REWARD FUNCTIONS ====================
@app.post("/reward-functions", response_model=schemas.RewardFunction, status_code=201)
def create_reward_function(
    payload: schemas.RewardFunctionCreate, db: Session = Depends(get_db)
):
    """Create a new reward function."""
    return crud.create_entity(db, models.RewardFunction, payload)


@app.get("/reward-functions", response_model=List[schemas.RewardFunction])
def list_reward_functions(db: Session = Depends(get_db)):
    """List all reward functions."""
    return crud.list_entities(db, models.RewardFunction)


@app.get(
    "/reward-functions/{reward_function_id}", response_model=schemas.RewardFunction
)
def get_reward_function(reward_function_id: int, db: Session = Depends(get_db)):
    """Get a specific reward function by ID."""
    return _get_or_404(db, models.RewardFunction, reward_function_id)


@app.put(
    "/reward-functions/{reward_function_id}", response_model=schemas.RewardFunction
)
def update_reward_function(
    reward_function_id: int,
    payload: schemas.RewardFunctionUpdate,
    db: Session = Depends(get_db),
):
    """Update a reward function. Fails if reward function is locked."""
    reward_fn = _get_or_404(db, models.RewardFunction, reward_function_id)
    crud.check_entity_locked(reward_fn)
    return crud.update_entity(db, reward_fn, payload)


@app.delete("/reward-functions/{reward_function_id}")
def delete_reward_function(reward_function_id: int, db: Session = Depends(get_db)):
    """Delete a reward function. Fails if used in any experiment."""
    reward_fn = _get_or_404(db, models.RewardFunction, reward_function_id)
    crud.check_experiments_using_entity(db, 'reward_function', reward_function_id)
    crud.delete_entity(db, reward_fn)
    return {"status": "deleted"}


# ==================== EXPERIMENTS ====================
@app.post("/experiments", response_model=schemas.Experiment, status_code=201)
def create_experiment(
    payload: schemas.ExperimentCreate, db: Session = Depends(get_db)
):
    """
    Create a new experiment and lock all related entities.
    This prevents modification of environments, agents, and reward functions.
    """
    # Verify all related entities exist
    _get_or_404(db, models.Environment, payload.env_id)
    _get_or_404(db, models.Agent, payload.agent_id)
    _get_or_404(db, models.RewardFunction, payload.reward_id)
    
    # Create experiment and lock entities
    return crud.create_experiment_and_lock_entities(db, payload.dict())


@app.get("/experiments", response_model=List[schemas.ExperimentSummary])
def list_experiments(db: Session = Depends(get_db)):
    """List all experiments with optimized loading (excludes large JSON fields)."""
    from sqlalchemy.orm import load_only, noload
    
    return db.query(models.Experiment).options(
        # Only load specific columns needed for list view (excludes trajectory_data which can be 100MB+)
        load_only(
            models.Experiment.id,
            models.Experiment.name,
            models.Experiment.type,
            models.Experiment.env_id,
            models.Experiment.agent_id,
            models.Experiment.reward_id,
            models.Experiment.model_snapshot_id,
            models.Experiment.base_model_snapshot_id,
            models.Experiment.fine_tuning_strategy,
            models.Experiment.training_mode,
            models.Experiment.residual_base_model_id,
            models.Experiment.residual_connector_id,
            models.Experiment.status,
            models.Experiment.total_steps,
            models.Experiment.current_step,
            models.Experiment.snapshot_freq,
            models.Experiment.max_ep_length,
            models.Experiment.evaluation_episodes,
            models.Experiment.fps_delay,
            models.Experiment.safety_constraint,  # Small JSON
            models.Experiment.seed,
            models.Experiment.created_at,
            # Explicitly NOT loading: trajectory_data (huge 80-200MB JSON)
        ),
        # Eager load required relationships
        joinedload(models.Experiment.environment),
        joinedload(models.Experiment.agent),
        joinedload(models.Experiment.reward_function),
        joinedload(models.Experiment.base_model_snapshot),
        joinedload(models.Experiment.residual_base_model),
        joinedload(models.Experiment.residual_connector),
        
        # Prevent loading metrics and models collections
        noload(models.Experiment.metrics),
        noload(models.Experiment.models)
    ).all()


@app.get("/experiments/{experiment_id}", response_model=schemas.Experiment)
def get_experiment(experiment_id: int, db: Session = Depends(get_db)):
    """Get a specific experiment by ID."""
    experiment = db.query(models.Experiment).options(
        joinedload(models.Experiment.environment),
        joinedload(models.Experiment.agent),
        joinedload(models.Experiment.reward_function)
    ).filter(models.Experiment.id == experiment_id).first()
    if not experiment:
        raise HTTPException(status_code=404, detail=f"Experiment {experiment_id} not found")
    
    # Add runtime status information
    experiment.is_running = experiment_runner.is_experiment_running(experiment_id)
    experiment.is_paused = experiment_runner.is_experiment_paused(experiment_id)
    
    return experiment


@app.put("/experiments/{experiment_id}", response_model=schemas.Experiment)
def update_experiment(
    experiment_id: int, payload: schemas.ExperimentUpdate, db: Session = Depends(get_db)
):
    """Update an experiment (name, status, parameters)."""
    experiment = _get_or_404(db, models.Experiment, experiment_id)
    return crud.update_entity(db, experiment, payload)


@app.delete("/experiments/{experiment_id}")
async def delete_experiment(experiment_id: int, db: Session = Depends(get_db)):
    """Delete an experiment and all related metrics/snapshots."""
    experiment = _get_or_404(db, models.Experiment, experiment_id)
    
    # Cancel experiment if it's still running
    if experiment_id in experiment_runner.running_tasks:
        print(f"[Delete] Cancelling running experiment {experiment_id}")
        await experiment_runner.cancel_experiment(experiment_id)
    
    # Clean up any lingering state in experiment runner
    if experiment_id in experiment_runner.cancel_events:
        del experiment_runner.cancel_events[experiment_id]
    if experiment_id in experiment_runner.paused_experiments:
        del experiment_runner.paused_experiments[experiment_id]
    if experiment_id in experiment_runner.models_cache:
        del experiment_runner.models_cache[experiment_id]
    if experiment_id in experiment_runner.envs_cache:
        try:
            experiment_runner.envs_cache[experiment_id].close()
        except:
            pass
        del experiment_runner.envs_cache[experiment_id]
    
    # Delete all model files associated with this experiment
    snapshots = db.query(models.ModelSnapshot).filter(
        models.ModelSnapshot.experiment_id == experiment_id
    ).all()
    
    for snapshot in snapshots:
        if snapshot.file_path and os.path.exists(snapshot.file_path):
            try:
                os.remove(snapshot.file_path)
                print(f"[Delete] Removed model file: {snapshot.file_path}")
            except Exception as e:
                print(f"[Delete] Failed to remove model file {snapshot.file_path}: {e}")
    
    # Delete experiment (cascade will handle snapshots and metrics in DB)
    crud.delete_entity(db, experiment)
    return {"status": "deleted"}


# ==================== MODEL SNAPSHOTS ====================
@app.post("/model-snapshots", response_model=schemas.ModelSnapshot, status_code=201)
def create_model_snapshot(
    payload: schemas.ModelSnapshotCreate, db: Session = Depends(get_db)
):
    """Create a new model snapshot for an experiment."""
    return crud.create_model_snapshot(
        db,
        experiment_id=payload.experiment_id,
        iteration=payload.iteration,
        file_path=payload.file_path,
        metrics=payload.metrics_at_save
    )


@app.get("/experiments/{experiment_id}/snapshots", response_model=List[schemas.ModelSnapshot])
def list_experiment_snapshots(experiment_id: int, db: Session = Depends(get_db)):
    """Get all model snapshots for an experiment."""
    _get_or_404(db, models.Experiment, experiment_id)
    return crud.get_snapshots_for_experiment(db, experiment_id)


@app.get("/snapshots", response_model=List[schemas.ModelSnapshot])
def list_all_snapshots(db: Session = Depends(get_db)):
    """Get all model snapshots across all experiments."""
    return db.query(models.ModelSnapshot).order_by(models.ModelSnapshot.created_at.desc()).all()


@app.delete("/snapshots/{snapshot_id}")
def delete_snapshot(snapshot_id: int, db: Session = Depends(get_db)):
    """Delete a model snapshot and its associated file."""
    snapshot = _get_or_404(db, models.ModelSnapshot, snapshot_id)
    
    # Delete the file from disk if it exists
    if snapshot.file_path and os.path.exists(snapshot.file_path):
        try:
            os.remove(snapshot.file_path)
            print(f"Deleted file: {snapshot.file_path}")
        except Exception as e:
            print(f"Failed to delete file {snapshot.file_path}: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to delete model file: {str(e)}")
    
    # Delete the database record
    crud.delete_entity(db, snapshot)
    
    return {"message": "Snapshot deleted successfully", "snapshot_id": snapshot_id}


@app.post("/snapshots/bulk-delete")
def bulk_delete_snapshots(snapshot_ids: List[int], db: Session = Depends(get_db)):
    """Delete multiple model snapshots and their associated files."""
    deleted_ids = []
    failed_ids = []
    
    for snapshot_id in snapshot_ids:
        snapshot = db.query(models.ModelSnapshot).filter(
            models.ModelSnapshot.id == snapshot_id
        ).first()
        
        if not snapshot:
            failed_ids.append({"id": snapshot_id, "error": "Not found"})
            continue
        
        # Delete the file from disk if it exists
        if snapshot.file_path and os.path.exists(snapshot.file_path):
            try:
                os.remove(snapshot.file_path)
                print(f"Deleted file: {snapshot.file_path}")
            except Exception as e:
                print(f"Failed to delete file {snapshot.file_path}: {e}")
                failed_ids.append({"id": snapshot_id, "error": f"Failed to delete file: {str(e)}"})
                continue
        
        # Delete the database record
        crud.delete_entity(db, snapshot)
        deleted_ids.append(snapshot_id)
    
    return {
        "message": f"Deleted {len(deleted_ids)} snapshots",
        "deleted_ids": deleted_ids,
        "failed_ids": failed_ids
    }


# ==================== EXPERIMENT METRICS ====================
@app.post("/experiment-metrics", response_model=schemas.ExperimentMetric, status_code=201)
def create_experiment_metric(
    payload: schemas.ExperimentMetricCreate, db: Session = Depends(get_db)
):
    """Create a new metric entry for an experiment."""
    return crud.create_metric(
        db,
        experiment_id=payload.experiment_id,
        step=payload.step,
        values=payload.values,
        logs=payload.logs
    )


@app.get("/experiments/{experiment_id}/metrics", response_model=List[schemas.ExperimentMetric])
def list_experiment_metrics(
    experiment_id: int, 
    limit: int = None, 
    db: Session = Depends(get_db)
):
    """Get metrics for an experiment, optionally limited."""
    _get_or_404(db, models.Experiment, experiment_id)
    return crud.get_metrics_for_experiment(db, experiment_id, limit)


@app.post("/experiments/bulk-latest-metrics")
def get_bulk_latest_metrics(
    experiment_ids: List[int],
    db: Session = Depends(get_db)
):
    """
    Get the latest metrics for multiple experiments in a single request.
    Optimized for bulk fetching to avoid N+1 query problems.
    
    Returns a dict mapping experiment_id -> latest metric data.
    """
    metrics_map = crud.get_bulk_latest_metrics(db, experiment_ids)
    return {
        exp_id: {
            "id": m.id,
            "experiment_id": m.experiment_id,
            "step": m.step,
            "values": m.values,
            "created_at": str(m.created_at)
        }
        for exp_id, m in metrics_map.items()
    }


# ==================== TRAJECTORY & REPLAY ====================
@app.get("/experiments/{experiment_id}/trajectory")
def get_experiment_trajectory(experiment_id: int, db: Session = Depends(get_db)):
    """Get the stored trajectory data for training or simulation experiment."""
    experiment = _get_or_404(db, models.Experiment, experiment_id)
    
    # Parse trajectory data if it exists
    trajectory = []
    total_frames = 0
    
    if experiment.trajectory_data:
        try:
            import json
            if isinstance(experiment.trajectory_data, str):
                trajectory = json.loads(experiment.trajectory_data)
            else:
                trajectory = experiment.trajectory_data
            total_frames = len(trajectory)
        except (json.JSONDecodeError, TypeError) as e:
            print(f"[API] Error parsing trajectory data for experiment {experiment_id}: {e}")
            trajectory = []
            total_frames = 0
    
    return {
        "experiment_id": experiment_id,
        "trajectory": trajectory,
        "total_frames": total_frames,
        "status": experiment.status,
        "type": experiment.type
    }


@app.get("/experiments/{experiment_id}/results")
def get_experiment_results(experiment_id: int, db: Session = Depends(get_db)):
    """Get evaluation results summary for completed experiments."""
    experiment = _get_or_404(db, models.Experiment, experiment_id)
    
    # Get recent metrics to calculate evaluation results
    metrics = crud.get_metrics_for_experiment(db, experiment_id, limit=1000)
    
    if not metrics:
        return {
            "experiment_id": experiment_id,
            "status": experiment.status,
            "has_results": False,
            "message": "No evaluation results available"
        }
    
    # Calculate evaluation statistics from metrics
    rewards = [m.values.get('total_reward', m.values.get('reward', 0)) for m in metrics if 'reward' in m.values or 'total_reward' in m.values]
    successes = [m.values.get('success', False) for m in metrics if 'success' in m.values]
    
    if rewards:
        avg_reward = sum(rewards) / len(rewards)
        success_rate = sum(1 for s in successes if s) / len(successes) if successes else 0.0
        
        results = {
            "experiment_id": experiment_id,
            "status": experiment.status,
            "has_results": True,
            "avg_reward": round(avg_reward, 2),
            "success_rate": round(success_rate * 100, 1),  # Convert to percentage
            "total_episodes": len(set([m.step // 1000 for m in metrics if m.step])) if metrics else 0,
            "total_steps": max([m.step for m in metrics if m.step]) if metrics else 0,
            "trajectory_frames": len(experiment.trajectory_data) if experiment.trajectory_data else 0
        }
    else:
        results = {
            "experiment_id": experiment_id,
            "status": experiment.status,
            "has_results": False,
            "message": "No evaluation metrics found"
        }
    
    return results




# ==================== EXPERIMENT CONTROL ====================
@app.post("/experiments/{experiment_id}/start")
async def start_experiment(experiment_id: int, db: Session = Depends(get_db)):
    """Start a training, simulation, or evaluation experiment."""
    experiment = _get_or_404(db, models.Experiment, experiment_id)
    
    print(f"\n>>> Received request to start experiment {experiment_id}")
    print(f">>> Experiment type: {experiment.type}")
    print(f">>> Experiment name: {experiment.name}")
    
    try:
        if experiment.type == "Training" or experiment.type == "Fine-Tuning":
            await experiment_runner.start_training(
                experiment, db, connection_manager.broadcast
            )
        elif experiment.type == "Simulation":
            await experiment_runner.start_simulation(
                experiment, db, connection_manager.broadcast
            )
        elif experiment.type == "Evaluation":
            await experiment_runner.start_evaluation(
                experiment, db, connection_manager.broadcast
            )
        else:
            raise HTTPException(status_code=400, detail="Invalid experiment type")
        
        return {"status": "success", "message": f"Experiment {experiment_id} started"}
    except RuntimeError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/experiments/{experiment_id}/pause")
def pause_experiment(experiment_id: int, db: Session = Depends(get_db)):
    """Pause a running experiment."""
    experiment = _get_or_404(db, models.Experiment, experiment_id)
    
    try:
        experiment_runner.pause_experiment(experiment_id)
        crud.update_experiment_status(db, experiment_id, "Paused")
        return {"status": "success", "message": f"Experiment {experiment_id} paused"}
    except RuntimeError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/experiments/{experiment_id}/resume")
async def resume_experiment(experiment_id: int, db: Session = Depends(get_db)):
    """Resume a paused experiment."""
    experiment = _get_or_404(db, models.Experiment, experiment_id)
    
    try:
        experiment_runner.resume_experiment(experiment_id)
        crud.update_experiment_status(db, experiment_id, "In Progress")
        return {"status": "success", "message": f"Experiment {experiment_id} resumed"}
    except RuntimeError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/experiments/{experiment_id}/cancel")
async def cancel_experiment(experiment_id: int, db: Session = Depends(get_db)):
    """Cancel a running or paused experiment."""
    experiment = _get_or_404(db, models.Experiment, experiment_id)
    
    # Check if experiment is part of a curriculum
    # if experiment.curriculum_step_id:
    #     raise HTTPException(
    #         status_code=400,
    #         detail="This experiment is part of a curriculum. Please cancel the curriculum instead."
    #     )
    
    # Check if actually running in background
    is_running_in_background = experiment_runner.is_experiment_running(experiment_id)
    
    # If experiment is Paused, update status directly
    if experiment.status == "Paused":
        experiment.status = "Cancelled"
        # Also clean up runner state if it thinks it's paused
        if experiment_runner.is_experiment_paused(experiment_id):
             experiment_runner.cleanup_experiment(experiment_id)
        db.commit()
        return {"status": "success", "message": f"Paused experiment {experiment_id} cancelled"}
    
    # If it's NOT running in background but status is In Progress, it's stuck
    # This handles cases where server restarted or job crashed without updating DB
    if experiment.status == "In Progress" and not is_running_in_background:
        print(f"[Cancel] Force cancelling stuck experiment {experiment_id}")
        experiment.status = "Cancelled"
        db.commit()
        # Ensure cleanup just in case (e.g. clear cache)
        experiment_runner.cleanup_experiment(experiment_id)
        return {"status": "success", "message": f"Stuck experiment {experiment_id} force cancelled"}
    
    # For running experiments, signal cancellation to the training thread
    try:
        await experiment_runner.cancel_experiment(experiment_id)
        # Update status immediately for simulations (training loop will update with correct step)
        if experiment.type == "Simulation":
            experiment.status = "Cancelled"
            db.commit()
            return {"status": "success", "message": f"Simulation experiment {experiment_id} cancelled"}
        # For training, let the loop handle status update with correct current_step
        return {"status": "success", "message": f"Experiment {experiment_id} cancellation requested"}
    except RuntimeError as e:
        # One last fallback: if cancellation failed with RuntimeError (maybe thread not found), 
        # but status is still In Progress, strictly force cancel
        if experiment.status == "In Progress":
             print(f"[Cancel] Runtime error during cancellation ({e}), force cancelling experiment {experiment_id}")
             experiment.status = "Cancelled"
             db.commit()
             experiment_runner.cleanup_experiment(experiment_id)
             return {"status": "success", "message": f"Experiment {experiment_id} force cancelled after error"}
             
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/experiments/{experiment_id}/export")
def export_experiment_report(experiment_id: int, db: Session = Depends(get_db)):
    """
    Export experiment report as Excel file with two sheets:
    1. Configuration: Agent, environment, reward, training settings, safety configs
    2. Metrics: Step-by-step metrics table
    """
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"Starting export for experiment {experiment_id}")
    
    # Load experiment with relationships (but NOT metrics - we'll load those separately)
    experiment = db.query(models.Experiment).options(
        joinedload(models.Experiment.environment),
        joinedload(models.Experiment.agent),
        joinedload(models.Experiment.reward_function),
        joinedload(models.Experiment.residual_connector)
    ).filter(models.Experiment.id == experiment_id).first()
    
    if not experiment:
        raise HTTPException(status_code=404, detail=f"Experiment {experiment_id} not found")
    
    logger.info(f"Experiment {experiment_id} loaded, creating workbook")
    
    # Create workbook
    wb = Workbook()
    
    # ==================== Sheet 1: Configuration ====================
    ws_config = wb.active
    ws_config.title = "Configuration"
    
    # Header styling
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Add experiment info
    row = 1
    ws_config.cell(row, 1, "Experiment Report").font = Font(bold=True, size=16)
    row += 2
    
    # Basic Information
    ws_config.cell(row, 1, "BASIC INFORMATION").font = header_font
    ws_config.cell(row, 1).fill = header_fill
    ws_config.merge_cells(f'A{row}:B{row}')
    row += 1
    
    info_data = [
        ("Experiment ID", experiment.id),
        ("Name", experiment.name),
        ("Type", experiment.type),
        ("Status", experiment.status),
        ("Created At", str(experiment.created_at)),
    ]
    
    for label, value in info_data:
        ws_config.cell(row, 1, label).font = Font(bold=True)
        ws_config.cell(row, 2, value)
        row += 1
    
    row += 1
    
    # Environment Configuration
    ws_config.cell(row, 1, "ENVIRONMENT").font = header_font
    ws_config.cell(row, 1).fill = header_fill
    ws_config.merge_cells(f'A{row}:B{row}')
    row += 1
    
    if experiment.environment:
        env_data = [
            ("Environment ID", experiment.env_id),
            ("Environment Name", experiment.environment.name),
            ("Length", experiment.environment.length),
            ("Width", experiment.environment.width),
            ("Height", experiment.environment.height),
            ("Buffer Start", experiment.environment.buffer_start),
            ("Buffer End", experiment.environment.buffer_end),
            ("Target Diameter", experiment.environment.target_diameter),
            ("Obstacles Count", len(experiment.environment.obstacles) if experiment.environment.obstacles else 0),
        ]
        for label, value in env_data:
            ws_config.cell(row, 1, label).font = Font(bold=True)
            ws_config.cell(row, 2, value)
            row += 1
    
    row += 1
    
    # Agent Configuration
    ws_config.cell(row, 1, "AGENT").font = header_font
    ws_config.cell(row, 1).fill = header_fill
    ws_config.merge_cells(f'A{row}:B{row}')
    row += 1
    
    if experiment.agent:
        agent_data = [
            ("Agent ID", experiment.agent_id),
            ("Agent Name", experiment.agent.name),
            ("Agent Type", experiment.agent.type),
            ("Agent Diameter", experiment.agent.agent_diameter),
            ("Max Speed", experiment.agent.agent_max_speed),
            ("Max Acceleration", experiment.agent.agent_max_acceleration),
            ("Lidar Range", experiment.agent.lidar_range),
            ("Lidar Rays", experiment.agent.lidar_rays),
            ("Kinematic Type", experiment.agent.kinematic_type),
        ]
        for label, value in agent_data:
            ws_config.cell(row, 1, label).font = Font(bold=True)
            ws_config.cell(row, 2, value)
            row += 1
        
        # Add agent hyperparameters if they exist
        if experiment.agent.parameters:
            ws_config.cell(row, 1, "Agent Hyperparameters:").font = Font(bold=True, italic=True)
            row += 1
            for param_key, param_value in experiment.agent.parameters.items():
                ws_config.cell(row, 1, f"  {param_key}").font = Font(italic=True)
                ws_config.cell(row, 2, str(param_value))
                row += 1
    
    row += 1
    
    # Reward Function
    ws_config.cell(row, 1, "REWARD FUNCTION").font = header_font
    ws_config.cell(row, 1).fill = header_fill
    ws_config.merge_cells(f'A{row}:B{row}')
    row += 1
    
    if experiment.reward_function:
        reward_data = [
            ("Reward Function ID", experiment.reward_id),
            ("Reward Function Name", experiment.reward_function.name),
            ("Weight Progress", experiment.reward_function.w_progress),
            ("Weight Time", experiment.reward_function.w_time),
            ("Weight Jerk", experiment.reward_function.w_jerk),
            ("Success Reward", experiment.reward_function.success_reward),
            ("Crash Reward", experiment.reward_function.crash_reward),
        ]
        for label, value in reward_data:
            ws_config.cell(row, 1, label).font = Font(bold=True)
            ws_config.cell(row, 2, value)
            row += 1
    
    row += 1
    
    # Training Configuration
    ws_config.cell(row, 1, "TRAINING CONFIGURATION").font = header_font
    ws_config.cell(row, 1).fill = header_fill
    ws_config.merge_cells(f'A{row}:B{row}')
    row += 1
    
    training_data = [
        ("Training Mode", experiment.training_mode or "Standard"),
        ("Total Steps", experiment.total_steps or "N/A"),
        ("Current Step", experiment.current_step or 0),
        ("Snapshot Frequency", experiment.snapshot_freq or "N/A"),
        ("Max Episode Length", experiment.max_ep_length or "N/A"),
    ]
    
    for label, value in training_data:
        ws_config.cell(row, 1, label).font = Font(bold=True)
        ws_config.cell(row, 2, value)
        row += 1
    
    row += 1
    
    # Residual Connector
    ws_config.cell(row, 1, "RESIDUAL LEARNING").font = header_font
    ws_config.cell(row, 1).fill = header_fill
    ws_config.merge_cells(f'A{row}:B{row}')
    row += 1
    
    if experiment.residual_connector:
        residual_data = [
            ("Residual Connector ID", experiment.residual_connector_id),
            ("Connector Name", experiment.residual_connector.name),
            ("Algorithm", experiment.residual_connector.algorithm),
            ("Parent Agent ID", experiment.residual_connector.parent_agent_id),
            ("Network Architecture", str(experiment.residual_connector.net_arch)),
        ]
        for label, value in residual_data:
            ws_config.cell(row, 1, label).font = Font(bold=True)
            ws_config.cell(row, 2, value)
            row += 1
        
        # Add residual connector hyperparameters if they exist
        if experiment.residual_connector.parameters:
            ws_config.cell(row, 1, "Residual Connector Hyperparameters:").font = Font(bold=True, italic=True)
            row += 1
            for param_key, param_value in experiment.residual_connector.parameters.items():
                ws_config.cell(row, 1, f"  {param_key}").font = Font(italic=True)
                ws_config.cell(row, 2, str(param_value))
                row += 1
        
        # Add base model info if in residual mode
        if experiment.training_mode == "residual" and experiment.residual_base_model_id:
            ws_config.cell(row, 1, "Residual Base Model ID").font = Font(bold=True)
            ws_config.cell(row, 2, experiment.residual_base_model_id)
            row += 1
    else:
        ws_config.cell(row, 1, "Residual Connector").font = Font(bold=True)
        ws_config.cell(row, 2, "Not Used")
        row += 1
    
    row += 1
    
    # Safety Constraints
    ws_config.cell(row, 1, "SAFETY CONSTRAINTS").font = header_font
    ws_config.cell(row, 1).fill = header_fill
    ws_config.merge_cells(f'A{row}:B{row}')
    row += 1
    
    safety_config = experiment.safety_constraint or {}
    if safety_config.get("enabled"):
        safety_data = [
            ("Safety Enabled", "Yes"),
            ("Risk Budget", safety_config.get("risk_budget", "N/A")),
            ("Initial Lambda", safety_config.get("initial_lambda", "N/A")),
            ("Lambda Learning Rate", safety_config.get("lambda_learning_rate", "N/A")),
            ("Update Frequency", safety_config.get("update_frequency", "N/A")),
            ("Cost Signal Preset", safety_config.get("cost_signal_preset", "N/A")),
            ("Collision Weight", safety_config.get("collision_weight", "N/A")),
            ("Near Miss Weight", safety_config.get("near_miss_weight", "N/A")),
            ("Danger Zone Weight", safety_config.get("danger_zone_weight", "N/A")),
            ("Near Miss Threshold", safety_config.get("near_miss_threshold", "N/A")),
            ("Ignore Walls", "Yes" if safety_config.get("ignore_walls", True) else "No"),
            ("Wall Near-Miss Weight", safety_config.get("wall_near_miss_weight", "N/A")),
            ("Wall Danger Zone Weight", safety_config.get("wall_danger_zone_weight", "N/A")),
            ("Wall Near-Miss Threshold", safety_config.get("wall_near_miss_threshold", "N/A")),
        ]
        for label, value in safety_data:
            ws_config.cell(row, 1, label).font = Font(bold=True)
            ws_config.cell(row, 2, value)
            row += 1
    else:
        ws_config.cell(row, 1, "Safety Enabled").font = Font(bold=True)
        ws_config.cell(row, 2, "No")
        row += 1
    
    # Auto-adjust column widths
    ws_config.column_dimensions['A'].width = 30
    ws_config.column_dimensions['B'].width = 50
    
    logger.info(f"Configuration sheet completed, loading metrics")
    
    # ==================== Sheet 2: Metrics ====================
    ws_metrics = wb.create_sheet("Metrics")
    
    # Load all metrics for this experiment
    metrics = db.query(models.ExperimentMetric).filter(
        models.ExperimentMetric.experiment_id == experiment_id
    ).order_by(models.ExperimentMetric.step).all()
    
    logger.info(f"Loaded {len(metrics)} metrics")
    
    if metrics:
        # Collect all unique metric names
        all_metric_names = set()
        for metric in metrics:
            if metric.values:
                all_metric_names.update(metric.values.keys())
        
        # Sort metric names for consistent column order
        metric_columns = ["step", "created_at"] + sorted(all_metric_names)
        
        # Write header row
        for col_idx, col_name in enumerate(metric_columns, 1):
            cell = ws_metrics.cell(1, col_idx, col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Write data rows
        for row_idx, metric in enumerate(metrics, 2):
            for col_idx, col_name in enumerate(metric_columns, 1):
                if col_name == "step":
                    ws_metrics.cell(row_idx, col_idx, metric.step)
                elif col_name == "created_at":
                    ws_metrics.cell(row_idx, col_idx, str(metric.created_at))
                else:
                    value = metric.values.get(col_name, "")
                    ws_metrics.cell(row_idx, col_idx, value)
        
        # Auto-adjust column widths using get_column_letter for proper Excel column names
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, len(metric_columns) + 1):
            ws_metrics.column_dimensions[get_column_letter(col_idx)].width = 15
    else:
        # No metrics yet
        ws_metrics.cell(1, 1, "No metrics data available yet").font = Font(italic=True)
    
    logger.info(f"Metrics sheet completed, saving workbook")
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    logger.info(f"Workbook saved, preparing response")
    
    # Create filename
    filename = f"experiment_{experiment_id}_{experiment.name.replace(' ', '_')}.xlsx"
    
    # Return as streaming response
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.post("/experiments/{experiment_id}/regenerate")
async def regenerate_simulation(experiment_id: int, db: Session = Depends(get_db)):
    """Regenerate simulation environment for a completed/cancelled/paused simulation experiment."""
    experiment = _get_or_404(db, models.Experiment, experiment_id)
    
    # Verify it's a simulation experiment
    if experiment.type != "Simulation":
        raise HTTPException(status_code=400, detail="Can only regenerate Simulation experiments")
    
    # Verify it's completed, cancelled, or paused
    if experiment.status not in ["Completed", "Cancelled", "Paused"]:
        raise HTTPException(status_code=400, detail="Can only regenerate completed, cancelled, or paused experiments")
    
    try:
        # Stop any running simulation first
        if experiment_id in experiment_runner.running_tasks:
            try:
                await experiment_runner.cancel_experiment(experiment_id)
                # Wait a bit for cleanup
                import asyncio
                await asyncio.sleep(0.5)
            except:
                pass
        
        # Clear the cache to force environment regeneration
        experiment_runner.clear_env_cache(experiment_id)
        
        # Reset experiment status to allow re-running
        experiment.status = "Paused"
        experiment.current_step = 0
        experiment.trajectory_data = None
        db.commit()
        
        # Refresh the experiment object from database
        db.refresh(experiment)
        
        # Auto-start the simulation to show new environment
        await experiment_runner.start_simulation(
            experiment, db, connection_manager.broadcast
        )
        
        return {"status": "success", "message": f"Environment regenerated and started for experiment {experiment_id}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to regenerate: {str(e)}")


# ==================== WEBSOCKET ====================
@app.websocket("/experiments/{experiment_id}/ws")
async def experiment_websocket(websocket: WebSocket, experiment_id: int):
    """WebSocket endpoint for real-time experiment updates."""
    await connection_manager.connect(websocket, experiment_id)
    
    # Send current experiment state immediately upon connection
    try:
        db = SessionLocal()
        experiment = crud.get_entity(db, models.Experiment, experiment_id)
        if experiment:
            # Send current experiment status
            current_state = {
                "type": "connection_status",
                "data": {
                    "experiment_id": experiment_id,
                    "status": experiment.status,
                    "type": experiment.type,
                    "current_step": getattr(experiment, 'current_step', 0),
                    "connected": True,
                    "timestamp": time.time()
                }
            }
            await websocket.send_json(current_state)
            print(f"[WebSocket] Sent connection_status for experiment {experiment_id}")
            
            # If experiment is running, send current progress
            if experiment_runner.is_experiment_running(experiment_id):
                # Get current step/episode info from database
                recent_metrics = crud.get_metrics_for_experiment(db, experiment_id, limit=10)
                current_episode = 1
                current_step = getattr(experiment, 'current_step', 0)
                total_episodes = getattr(experiment, 'evaluation_episodes', 100) if experiment.type == 'evaluation' else 1
                
                # Try to determine current episode from recent metrics
                if recent_metrics and experiment.type == 'evaluation':
                    # Look for episode completion patterns in recent metrics
                    for metric in recent_metrics[::-1]:  # Check most recent first
                        if 'episode' in metric.logs.lower() or 'complete' in metric.logs.lower():
                            try:
                                import re
                                episode_match = re.search(r'Episode (\d+)', metric.logs)
                                if episode_match:
                                    current_episode = int(episode_match.group(1))
                                    break
                            except:
                                pass
                
                progress_state = {
                    "type": "progress", 
                    "data": {
                        "experiment_id": experiment_id,
                        "step": current_step,
                        "total_steps": 0,  # Not used for evaluation
                        "episode": current_episode,
                        "total_episodes": total_episodes,
                        "metrics": {},  # Empty metrics object
                        "status": "In Progress",
                        "timestamp": time.time()
                    }
                }
                await websocket.send_json(progress_state)
                print(f"[WebSocket] Sent progress state for running experiment {experiment_id}: episode {current_episode}/{total_episodes}")
            
            elif experiment.status in ['Completed', 'Cancelled'] and experiment.type == 'evaluation':
                # For completed evaluations, send final results with correct episode count
                all_metrics = crud.get_metrics_for_experiment(db, experiment_id, limit=1000)
                total_episodes = getattr(experiment, 'evaluation_episodes', 100)
                
                # Count actual completed episodes from metrics
                completed_episodes = len(all_metrics) if all_metrics else total_episodes
                
                # Calculate final metrics
                rewards = [m.values.get('total_reward', m.values.get('reward', 0)) for m in all_metrics if 'reward' in m.values or 'total_reward' in m.values]
                successes = [m.values.get('success', False) for m in all_metrics if 'success' in m.values]
                
                final_metrics = {}
                if rewards:
                    final_metrics = {
                        'avg_reward': sum(rewards) / len(rewards),
                        'success_rate': sum(1 for s in successes if s) / len(successes) if successes else 0.0,
                    }
                
                completion_state = {
                    "type": "progress",
                    "data": {
                        "experiment_id": experiment_id,
                        "step": 0,
                        "total_steps": 0,
                        "episode": completed_episodes,
                        "total_episodes": total_episodes,
                        "metrics": final_metrics,
                        "status": experiment.status,
                        "timestamp": time.time()
                    }
                }
                await websocket.send_json(completion_state)
                print(f"[WebSocket] Sent completion state for {experiment.status.lower()} evaluation {experiment_id}: {completed_episodes}/{total_episodes} episodes")
            elif experiment.status == "Completed" and experiment.type == "evaluation":
                # Send evaluation results for completed evaluations
                try:
                    metrics = crud.get_metrics_for_experiment(db, experiment_id, limit=500)
                    if metrics:
                        rewards = [m.values.get('total_reward', m.values.get('reward', 0)) for m in metrics if 'reward' in m.values or 'total_reward' in m.values]
                        successes = [m.values.get('success', False) for m in metrics if 'success' in m.values]
                        
                        if rewards:
                            avg_reward = sum(rewards) / len(rewards)
                            success_rate = sum(1 for s in successes if s) / len(successes) if successes else 0.0
                            
                            results_state = {
                                "type": "progress",
                                "data": {
                                    "experiment_id": experiment_id,
                                    "status": "Completed",
                                    "episode": len(rewards),
                                    "total_episodes": len(rewards),
                                    "completed": True,
                                    "avg_reward": round(avg_reward, 2),
                                    "success_rate": round(success_rate, 3),
                                    "success_rate_percent": round(success_rate * 100, 1),
                                    "total_frames": len(experiment.trajectory_data) if experiment.trajectory_data else 0,
                                    "has_trajectory": bool(experiment.trajectory_data),
                                    "progress_percent": 100.0,
                                    "timestamp": time.time()
                                }
                            }
                            await websocket.send_json(results_state)
                            print(f"[WebSocket] Sent evaluation results for experiment {experiment_id}: {len(rewards)} episodes")
                except Exception as e:
                    print(f"[WebSocket] Error sending evaluation results: {e}")
        
        db.close()
    except Exception as e:
        print(f"[WebSocket] Error sending initial state: {e}")
    
    try:
        while True:
            # Keep connection alive, actual broadcasting happens via connection_manager
            await websocket.receive_text()
    except WebSocketDisconnect:
        connection_manager.disconnect(websocket, experiment_id)





# ==================== TELEMETRY ====================
async def telemetry_stream(websocket: WebSocket) -> None:
    """Stream telemetry data to connected clients."""
    await websocket.accept()
    try:
        while True:
            now = time.time()
            payload = schemas.TelemetryMessage(
                timestamp=now,
                position=(0.0, 0.0, 1.0),
                orientation=(0.0, 0.0, 0.0),
                linear_velocity=(0.0, 0.0, 0.0),
                angular_velocity=(0.0, 0.0, 0.0),
            )
            await websocket.send_json(payload.dict())
            await asyncio.sleep(0.5)
    except WebSocketDisconnect:
        return


@app.websocket("/ws/telemetry")
async def telemetry_ws(websocket: WebSocket):
    """WebSocket endpoint for drone telemetry."""
    await telemetry_stream(websocket)


# ==================== CURRICULUMS ====================
# @app.post("/curriculums", response_model=schemas.CurriculumRead, status_code=201)
# def create_curriculum(payload: schemas.CurriculumCreate, db: Session = Depends(get_db)):
#     """Create a new curriculum pipeline."""
#     # Create curriculum
#     curriculum = models.Curriculum(
#         name=payload.name,
#         description=payload.description,
#         status=models.CurriculumStatus.PENDING,
#         current_step_order=1
#     )
#     db.add(curriculum)
#     db.flush()
    
#     # Create steps
#     for step_data in payload.steps:
#         step_dict = step_data.dict()
#         step = models.CurriculumStep(
#             curriculum_id=curriculum.id,
#             order=step_dict["order"],
#             step_type=step_dict["step_type"],
#             config=step_dict["config"],
#             safety_config=step_dict["safety_config"],
#             target_success_rate=step_dict["target_success_rate"],
#             min_episodes=step_dict["min_episodes"],
#             is_completed=False
#         )
#         db.add(step)
    
#     db.commit()
#     db.refresh(curriculum)
#     return curriculum


# @app.get("/curriculums", response_model=List[schemas.CurriculumRead])
# def list_curriculums(db: Session = Depends(get_db)):
#     """List all curriculum pipelines."""
#     return db.query(models.Curriculum).options(
#         joinedload(models.Curriculum.steps)
#     ).all()


# @app.get("/curriculums/{curriculum_id}", response_model=schemas.CurriculumRead)
# def get_curriculum(curriculum_id: int, db: Session = Depends(get_db)):
#     """Get a specific curriculum by ID."""
#     curriculum = db.query(models.Curriculum).options(
#         joinedload(models.Curriculum.steps)
#     ).filter(models.Curriculum.id == curriculum_id).first()
#     if not curriculum:
#         raise HTTPException(status_code=404, detail="Curriculum not found")
#     return curriculum


# @app.post("/curriculums/{curriculum_id}/start")
# def start_curriculum(curriculum_id: int, db: Session = Depends(get_db)):
#     """Start a curriculum pipeline."""
#     curriculum = _get_or_404(db, models.Curriculum, curriculum_id)
    
#     if curriculum.status == models.CurriculumStatus.RUNNING:
#         raise HTTPException(status_code=400, detail="Curriculum is already running")
    
#     curriculum.status = models.CurriculumStatus.RUNNING
#     db.commit()
#     return {"status": "success", "message": f"Curriculum {curriculum_id} started"}


# @app.post("/curriculums/{curriculum_id}/pause")
# def pause_curriculum(curriculum_id: int, db: Session = Depends(get_db)):
#     """Pause a running curriculum pipeline."""
#     curriculum = _get_or_404(db, models.Curriculum, curriculum_id)
    
#     if curriculum.status != models.CurriculumStatus.RUNNING:
#         raise HTTPException(status_code=400, detail="Curriculum is not running")
    
#     curriculum.status = models.CurriculumStatus.PAUSED
#     db.commit()
    
#     # Also pause the current experiment if running
#     current_step = db.query(models.CurriculumStep).filter(
#         models.CurriculumStep.curriculum_id == curriculum_id,
#         models.CurriculumStep.order == curriculum.current_step_order
#     ).first()
    
#     if current_step and current_step.experiment:
#         try:
#             experiment_runner.pause_experiment(current_step.experiment.id)
#             crud.update_experiment_status(db, current_step.experiment.id, "Paused")
#         except Exception as e:
#             print(f"Warning: Failed to pause experiment {current_step.experiment.id}: {e}")
    
#     return {"status": "success", "message": f"Curriculum {curriculum_id} paused"}


# @app.post("/curriculums/{curriculum_id}/cancel")
# def cancel_curriculum(curriculum_id: int, db: Session = Depends(get_db)):
#     """Cancel a running curriculum pipeline and its active experiment."""
#     curriculum = _get_or_404(db, models.Curriculum, curriculum_id)
    
#     if curriculum.status not in [models.CurriculumStatus.RUNNING, models.CurriculumStatus.PAUSED]:
#         raise HTTPException(status_code=400, detail="Curriculum is not running or paused")
    
#     # Cancel ALL experiments associated with this curriculum's steps
#     steps = db.query(models.CurriculumStep).filter(
#         models.CurriculumStep.curriculum_id == curriculum_id
#     ).all()
    
#     for step in steps:
#         if step.experiment:
#             try:
#                 # Unlink experiment from curriculum first so it can be cancelled independently
#                 experiment = step.experiment
#                 experiment.curriculum_step_id = None
#                 db.add(experiment)
#                 db.flush()
                
#                 if experiment.status == models.ExperimentStatus.IN_PROGRESS:
#                     # Note: This is synchronous, ideally should be async but we can't await here easily without async def
#                     # Ideally user should use the async endpoint logic or we convert this to async
#                     # For now, just mark as cancelled in DB
#                     experiment.status = "Cancelled"
#                     print(f"Marked experiment {experiment.id} as Cancelled for curriculum {curriculum_id}")
#             except Exception as e:
#                 print(f"Warning: Failed to cancel experiment {step.experiment.id}: {e}")
    
#     # Mark curriculum as failed (cancelled)
#     curriculum.status = models.CurriculumStatus.FAILED
#     db.commit()
    
#     return {"status": "success", "message": f"Curriculum {curriculum_id} cancelled"}


# @app.delete("/curriculums/{curriculum_id}")
# def delete_curriculum(curriculum_id: int, db: Session = Depends(get_db)):
#     """Delete a curriculum."""
#     curriculum = _get_or_404(db, models.Curriculum, curriculum_id)
    
#     # Cancel and unlink any running experiments first
#     steps = db.query(models.CurriculumStep).filter(
#         models.CurriculumStep.curriculum_id == curriculum_id
#     ).all()
    
#     for step in steps:
#         if step.experiment:
#             try:
#                 # Unlink experiment from curriculum
#                 experiment = step.experiment
#                 experiment.curriculum_step_id = None
#                 db.add(experiment)
                
#                 # Cancel if still running
#                 if experiment.status == models.ExperimentStatus.IN_PROGRESS:
#                     experiment.status = "Cancelled"
#                     print(f"Marked experiment {experiment.id} as Cancelled before deleting curriculum {curriculum_id}")
#             except Exception as e:
#                 print(f"Warning: Failed to handle experiment {step.experiment.id}: {e}")
    
#     db.flush()
    
#     # Now delete the curriculum
#     crud.delete_entity(db, curriculum)
#     return {"status": "deleted"}


@app.get("/job-capacity")
def get_job_capacity():
    """Get information about background job capacity."""
    return experiment_runner.get_job_capacity_info()


@app.get("/dashboard/stats")
def get_dashboard_stats(db: Session = Depends(get_db)):
    """Get dashboard statistics."""
    from sqlalchemy import func
    
    # Count total entities
    total_environments = db.query(func.count(models.Environment.id)).scalar()
    total_agents = db.query(func.count(models.Agent.id)).scalar()
    total_reward_functions = db.query(func.count(models.RewardFunction.id)).scalar()
    total_experiments = db.query(func.count(models.Experiment.id)).scalar()
    total_models = db.query(func.count(models.ModelSnapshot.id)).scalar()
    # total_curriculums = db.query(func.count(models.Curriculum.id)).scalar()
    total_residual_connectors = db.query(func.count(models.ResidualConnector.id)).scalar()
    
    # Count experiments by type
    experiments_by_type = db.query(
        models.Experiment.type,
        func.count(models.Experiment.id)
    ).group_by(models.Experiment.type).all()
    
    experiment_type_counts = {
        exp_type.value: count for exp_type, count in experiments_by_type
    }
    
    # Count experiments by status
    experiments_by_status = db.query(
        models.Experiment.status,
        func.count(models.Experiment.id)
    ).group_by(models.Experiment.status).all()
    
    experiment_status_counts = {
        status.value: count for status, count in experiments_by_status
    }
    
    # # Count curriculums by status
    # curriculums_by_status = db.query(
    #     models.Curriculum.status,
    #     func.count(models.Curriculum.id)
    # ).group_by(models.Curriculum.status).all()
    
    # curriculum_status_counts = {
    #     status.value: count for status, count in curriculums_by_status
    # }
    
    return {
        "totals": {
            "environments": total_environments,
            "agents": total_agents,
            "reward_functions": total_reward_functions,
            "experiments": total_experiments,
            "models": total_models,
            "curriculums": 0, # total_curriculums,
            "residual_connectors": total_residual_connectors,
        },
        "experiments_by_type": experiment_type_counts,
        "experiments_by_status": experiment_status_counts,
        "curriculums_by_status": {}, # curriculum_status_counts,
    }


@app.get("/")
def root():
    """Root endpoint - health check."""
    return {"message": "ConflictAwareLab API running", "version": "0.1.0"}
